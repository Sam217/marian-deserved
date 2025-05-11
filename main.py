import csv
import sqlite3
import sys
import os
import unicodedata
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl.styles


GDataTypesCZECH = {
    "datum": "TEXT",
    "číslo": "INTEGER",
    "cen": "REAL",
    "celk": "REAL",
    "DPH": "REAL",
    "spár": "INTEGER",
    "tisk": "INTEGER",
    "potvrz": "INTEGER",
    "schvál": "INTEGER",
}

GSqlGoodsTypeStrs = {'Obleceni': 'oble', 'Boty': 'boty',
                     'Kosmetika': 'kosme', 'Kabelky': 'kabel'}

GCartonWeight = 1235e-6  # 1235 g = 1.235 kg
GgoodsCoefficients = {
    "Obleceni": {"plast": 13e-6, "papir": 0, "lepenka": 40},
    "Boty": {"plast": 0, "papir": 270e-6, "lepenka": 8},
    "Kosmetika": {"plast": 0, "papir": 0, "lepenka": 0},
    "Kabelky": {"plast": 139e-6, "papir": 0, "lepenka": 12},
}

GcoefficientsSimple = [
    ("Obleceni", 13e-6, 0, 40),
    ("Boty", 0, 270e-6, 8),
    ("Kosmetika", 0, 0, 0),
    ("Kabelky", 139e-6, 0, 12),
]

# Grequired_cols = ["Dodavatel", "Typ_zbozi", "Množství celkem"]


class GoodsType:
    def __init__(self, name, filterStr, plast=0, papir=0, lepenka=0):
        self.name = name
        self.filterStr = filterStr
        self.plast = plast
        self.papir = papir
        self.lepenka = lepenka

    def __repr__(self):
        return f"GoodsType('{self.name}', plast={self.plast}, papir={self.papir}, lepenka={self.lepenka})"

    def ToStrList(self):
        return [self.name, str(self.plast), str(self.papir), str(self.lepenka)]


# Convert the dictionary to a list of GoodsType instances
GgoodsList = [
    GoodsType("Obleceni", "oble", plast=13e-6, lepenka=40),
    GoodsType("Boty", "boty", papir=270e-6, lepenka=8),
    GoodsType("Kosmetika", "kosme"),
    GoodsType("Kabelky", "kabel", plast=139e-6, lepenka=12),
]

# Definovane typy zbozi
for goods in GgoodsList:
    print(
        f"Zbozi: {goods.name}: plast={goods.plast}, papir={goods.papir}, lepenka={goods.lepenka}")


def removeDiacritics(instr):
    # Normalize the string - remove diacritics
    n = unicodedata.normalize("NFKD", instr)
    res = "".join([c for c in n if not unicodedata.combining(c)])
    return res


def createDBoverwrite():
    # Step 1: Connect to SQLite database (creates one if it doesn't exist)
    dbName = "csvimported.db"
    if os.path.exists(dbName):
        print(f"{dbName} file exists, removing...")
        os.remove(dbName)

    conn = sqlite3.connect(dbName)
    return conn


def csvToSqlite(cursor, sourceCsv, suppliersCountryCsv):
    with open(suppliersCountryCsv, "r", encoding="utf8") as supplierList:
        # Step 2: Read the CSV file
        with open(sourceCsv, "r", encoding="utf8") as file:
            importedCSVreader = csv.reader(file)
            # Get column headers from first row
            headers = next(importedCSVreader)

            # read suppliers table
            supplierCsvReader = csv.reader(supplierList, delimiter=";")
            supplierheader = next(supplierCsvReader)

            supplierTableName = "suppliersCountry"
            columnsSup = []
            for s in supplierheader:
                s = s.replace(" ", "_")
                normalized = removeDiacritics(s)
                # DEBUG print
                print(f"{s} -> {normalized}")
                columnsSup.append(normalized)

            queryCreateSuppTable = f'CREATE TABLE IF NOT EXISTS {supplierTableName} ({", ".join(columnsSup)})'
            cursor.execute(queryCreateSuppTable)

            supplierValues = ", ".join(["?" for _ in supplierheader])
            # DEBUG print header
            # print(f"supplierValues: {supplierValues}")
            queryInsertSup = (
                f"INSERT INTO {supplierTableName} VALUES ({supplierValues})"
            )

            for row in supplierCsvReader:
                cursor.execute(queryInsertSup, row)

            sqliteDataTypes = []
            # process header data types
            for h in headers:
                typeFound = False
                for name, type in GDataTypesCZECH.items():
                    if h.lower().find(name) != -1:
                        sqliteDataTypes.append(type)
                        typeFound = True
                        break

                # everything that is not easily identifiable is TEXT
                if typeFound == False:
                    sqliteDataTypes.append("TEXT")

            # Step 3: Create table dynamically based on CSV headers
            # Replace spaces with underscores and handle special characters if needed
            columns = []
            for i, h in enumerate(headers):
                h1 = h.replace(" ", "_")
                normalized = removeDiacritics(h1)
                # DEBUG print
                # print(f"{h} -> {h1} -> {normalized}")
                columns.insert(i, f'"{normalized}" {sqliteDataTypes[i]}')

            productsTableName = "suppliedProducts"
            queryCreateTable = (
                f'CREATE TABLE IF NOT EXISTS {productsTableName} ({", ".join(columns)})'
            )

            # DEBUG print sqlite types
            # print(f"queryCreateTable: {queryCreateTable}")
            cursor.execute(queryCreateTable)

            # Step 4: Prepare INSERT query
            placeholders = ", ".join(["?" for _ in headers])
            queryInsert = f"INSERT INTO {productsTableName} VALUES ({placeholders})"

            # Step 5: Insert CSV data into the table
            for row in importedCSVreader:
                cursor.execute(queryInsert, row)


def createCoeffsTable(cursor, goodsTypeStr, coeffsTable):
    queryCreateCoeffsTable = f"CREATE TABLE IF NOT EXISTS {coeffsTable} ({goodsTypeStr}, koef_plast, koef_papir, koef_lepenka)"
    cursor.execute(queryCreateCoeffsTable)

    insertCoeffData = f"INSERT INTO {coeffsTable} VALUES (?,?,?,?)"

    for type in GgoodsList:
        # DEBUG print
        # print(f"coef data: {type.ToStrList()}")
        cursor.execute(insertCoeffData, type.ToStrList())


def calcViewTotals(sqlCursor, viewName, selectFrom):
    resultViewQuery = f"""
    CREATE VIEW IF NOT EXISTS {viewName} AS
        SELECT
            SUM(e.'Plast [g]') as 'Plasty celkem [g]',
			SUM(e.'Papir [g]') as 'Papir celkem [g]',
			SUM(e.'Lepenka [g]') as 'Lepenka celkem [g]'
		FROM {selectFrom} as e;
    """
    sqlCursor.execute(resultViewQuery)


def calcViewTotalsPerType(sqlCursor, viewName, selectFrom, typeFilterStr):
    resultViewQuery = f"""
    CREATE VIEW IF NOT EXISTS {viewName} AS
        SELECT
            SUM(e.'Plast [g]') as 'Plasty celkem [g]',
			SUM(e.'Papir [g]') as 'Papir celkem [g]',
			SUM(e.'Lepenka [g]') as 'Lepenka celkem [g]'
		FROM {selectFrom} as e
        WHERE e.Typ_zbozi LIKE '%{typeFilterStr}%';
    """
    sqlCursor.execute(resultViewQuery)


def createFilterByCountryView(sqlCursor, viewName, selectFrom, filterString):
    materialsViewQuery = f"""
    CREATE VIEW IF NOT EXISTS {viewName} AS
        SELECT
            *
		FROM {selectFrom} as e WHERE
			e.'PuvodCZ' LIKE '%{filterString}%';
    """
    sqlCursor.execute(materialsViewQuery)


def buildDB(sourceCsv, suppliersCountryCsv):
    conn = createDBoverwrite()
    cursor = conn.cursor()
    csvToSqlite(cursor, sourceCsv, suppliersCountryCsv)

    # DB postprocessing, data preparation
    totalOblec = 0
    totalBoty = 0
    totalKosme = 0
    totalKabel = 0

    # helper sql strings
    sqlQueryJoinCommonAll = "suppliedProducts as sp JOIN suppliersCountry as sc ON sp.Dodavatel LIKE '%' || sc.Dodavel || '%'"
    sqlQueryJoinCommonEU = "suppliedProducts as sp JOIN suppliersCountry as sc ON sp.Dodavatel LIKE '%' || sc.Dodavel || '%' WHERE sc._CZ_ano_ne LIKE '%ne%'"
    sqlQueryJoinCommonCZ = "suppliedProducts as sp JOIN suppliersCountry as sc ON sp.Dodavatel LIKE '%' || sc.Dodavel || '%' WHERE sc._CZ_ano_ne not LIKE '%ne%'"

    goodsTypeStr = "Typ_zbozi"
    goodsCountStr = "Mnozstvi_celkem"
    goodsViewName = "zbozi_puvod"

    # ================================================================================================= #
    # create coefficients table if not already present
    # ================================================================================================= #
    coeffsTable = "coefficients"
    createCoeffsTable(cursor, goodsTypeStr, coeffsTable)

    # ================================================================================================= #
    # create a view that joins the main table with the table containing suppliers & country (_CZ_ano_ne)
    # ================================================================================================= #
    joinQueryAll = f"SELECT * FROM {sqlQueryJoinCommonAll}"
    crateJoinedViewAll = f"""
    CREATE VIEW IF NOT EXISTS {goodsViewName} AS
    {joinQueryAll}
    """
    cursor.execute(crateJoinedViewAll)

    # ================================================================================================= #
    # create a view that selects from joined view & counts total amount
    # ================================================================================================= #
    goodsByTypeView = "zbozi_podle_typu"

    caseStr = '\r\n'.join(
        [f'WHEN {goodsTypeStr} LIKE \'%{type.filterStr}%\' THEN \'{type.name.lower()}\'' for type in GgoodsList])

    print(f'caseStr = {caseStr}')

    goodsByTypeViewQ = f"""
        CREATE VIEW IF NOT EXISTS {goodsByTypeView} AS
        SELECT
            {goodsTypeStr},
            Dodavatel,
            _CZ_ano_ne,
            {goodsCountStr},
            SUM({goodsCountStr}) as total_amount
        FROM
            {goodsViewName} as gv
        GROUP BY
            CASE
                {caseStr}
                ELSE NULL
            END,
            Dodavatel, _CZ_ano_ne
        """

    # the command after CASE should produce:
    # WHEN {goodsTypeStr} LIKE '%{GSqlGoodsTypeStrs['Obleceni']}%' THEN 'Obleceni''
    # WHEN {goodsTypeStr} LIKE '%{GSqlGoodsTypeStrs['Boty']}%' THEN 'Boty'
    # WHEN {goodsTypeStr} LIKE '%{GSqlGoodsTypeStrs['Kosmetika']}%' THEN 'Kosmetika'a'
    # WHEN {goodsTypeStr} LIKE '%{GSqlGoodsTypeStrs['Kabelky']}%' THEN 'Kabelky'

    # WHEN gv.{goodsTypeStr} LIKE '%oble%' THEN 'Obleceni'
    # WHEN gv.{goodsTypeStr} LIKE '%boty%' THEN 'Boty'
    # WHEN gv.{goodsTypeStr} LIKE '%kosme%' THEN 'Kosmetika'
    # WHEN gv.{goodsTypeStr} LIKE '%kabel%' THEN 'Kabelky'

    cursor.execute(goodsByTypeViewQ)

    sqlColumnsShared = "Dodavatel, Puvod, Mnozstvi, Papir, Plast, Lepenka"

    plasticPaperCartonView = "ekokom_res"

    plasticPaperCartonViewQ = f"""
    CREATE VIEW IF NOT EXISTS {plasticPaperCartonView} AS
        SELECT
            gv.Dodavatel,
            gv.{goodsTypeStr},
            gv.total_amount,
            gv._CZ_ano_ne as PuvodCZ,
            total_amount * c.koef_plast * 1E6 as 'Plast [g]',
            total_amount * c.koef_papir * 1E6 as 'Papir [g]',
            total_amount / c.koef_lepenka * {str(GCartonWeight)} * 1E6 as 'Lepenka [g]'
        FROM
            {goodsByTypeView} as gv
        JOIN
            {coeffsTable} as c ON  c.{goodsTypeStr} LIKE '%' ||
                CASE
                    WHEN gv.{goodsTypeStr} LIKE '%oble%' THEN 'obleceni'
                    WHEN gv.{goodsTypeStr} LIKE '%boty%' THEN 'boty'
                    WHEN gv.{goodsTypeStr} LIKE '%kosme%' THEN 'kosmetika'
                    WHEN gv.{goodsTypeStr} LIKE '%kabel%' THEN 'kabelky'
                    ELSE NULL
                END || '%'
    """
    cursor.execute(plasticPaperCartonViewQ)

    materialsCZview = "ekokom_CZ"
    materialsEU_USview = "ekokom_import"

    createFilterByCountryView(cursor, materialsCZview,
                              plasticPaperCartonView, "ano")
    createFilterByCountryView(cursor, materialsEU_USview,
                              plasticPaperCartonView, "ne")

    materialsCZviewObleceni = "ekokom_CZobleceni"
    materialsEU_USviewObleceni = "ekokom_importObleceni"

    # materialsCZviewTypes = [f"ekokom_CZ{type.name}" for type in GgoodsList]
    # materialsEU_USviewTypes = [
    #     f"ekokom_importObleceni{type.name}" for type in GgoodsList]
    materialsCZviewTypes = []
    materialsEU_USviewTypes = []

    for t in GgoodsList:
        calcViewTotalsPerType(
            cursor, f"ekokom_CZ{t.name}", materialsCZview, t.filterStr)
        materialsCZviewTypes.append(f"ekokom_CZ{t.name}")
        calcViewTotalsPerType(
            cursor, f"ekokom_import{t.name}", materialsEU_USview, t.filterStr)
        materialsEU_USviewTypes.append(f"ekokom_import{t.name}")

    resultCZview = "ekokom_totalCZ"
    resultEU_USview = "ekokom_totalImport"

    calcViewTotals(cursor, resultCZview, materialsCZview)
    calcViewTotals(cursor, resultEU_USview, materialsEU_USview)

    # PrintOutDemoResult(conn, cursor, totalOblec, totalBoty,
    #                    totalKosme, totalKabel, sqlQueryJoinCommonCZ, goodsTypeStr)

    print("CSV data successfully imported into SQLite database!")

    wb = Workbook()
    # ws = wb.active
    WriteToXLSX(cursor, materialsCZview,
                materialsCZviewTypes, resultCZview, wb)
    WriteToXLSX(cursor, materialsEU_USview,
                materialsEU_USviewTypes, resultEU_USview, wb)

    # Uložení souboru
    wb.save("ekokom.xlsx")

    # Step 6: Commit changes and close connection
    conn.commit()
    conn.close()


def WriteToXLSX(sqlCursor, materialsView, materialsViewTypes, resultView, wb):
    wb.create_sheet(materialsView)
    ws = wb[materialsView]

    # Zápis hlavičky
    ws.append(["Dodavatel", "Kategorie", "Množství", "PůvodCZ",
              'Plast [g]', 'Papir [g]', 'Lepenka [g]'])

    # Zápis dat
    xlsxEkokomCZquery = f"""
        SELECT * FROM {materialsView}
    """

    qResult = sqlCursor.execute(xlsxEkokomCZquery)
    numRows = 0
    for row in qResult.fetchall():
        ws.append(row)
        numRows += 1

    # tucne
    head = ws[1]
    boldFont = openpyxl.styles.Font(bold=True)
    sideStyle = openpyxl.styles.Side(style='thin')

    for cell in head:
        cell.font = boldFont

    # 4️⃣ — vytvoříme styl ohraničení
    thinBorder = openpyxl.styles.Border(
        left=sideStyle,
        right=sideStyle,
        top=sideStyle,
        bottom=sideStyle
    )

    # 5️⃣ — projdeme všechny buňky a přidáme ohraničení
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thinBorder

    ws.append([""])
    ws.append([""])

    columnCatnames = ["Kategorie celkem",
                      'Plast [g]', 'Papir [g]', 'Lepenka [g]']

    currentXlsxRow = numRows + 2  # skip two rows to add some space for readability
    currentXlsxColumn = 3

    # Zápis hlavičky ručně, se stylem
    for col_index, col_name in enumerate(columnCatnames, start=1):
        cell = ws.cell(row=currentXlsxRow,
                       column=col_index + currentXlsxColumn, value=col_name)
        cell.font = boldFont
        cell.border = thinBorder

    currentXlsxRow += 1

    for i, t in enumerate(GgoodsList):
        # print(f'attempting to export {materialsCZviewTypes[i]} to xlsx...')
        qResult = sqlCursor.execute(f"""
                                 SELECT * FROM {materialsViewTypes[i]}
                                 """)

        rowData = []
        rowData.append(f"{t.name}")
        for q in qResult.fetchall():
            for c in q:
                rowData.append(c)

        for col_index, value in enumerate(rowData, start=1):
            cell = ws.cell(row=i + currentXlsxRow,
                           column=col_index + currentXlsxColumn, value=value)
            if col_index == 0:
                cell.font = boldFont
            cell.border = thinBorder

        # rowData = []
        # rowData.append("")
        # rowData.append("")
        # rowData.append("")
        # rowData.append(f"{t.name}")
        # for q in qResult.fetchall():
        #     for c in q:
        #         rowData.append(c)

        # print(f"rowData = {rowData}")
        # ws.append(rowData)

    ws.append([""])
    qResult = sqlCursor.execute(f"""
                             SELECT * FROM {resultView}
                             """)

    rowData = []
    rowData.append("CELKEM")
    for q in qResult.fetchall():
        for c in q:
            rowData.append(c)

    for col_index, value in enumerate(rowData, start=1):
        cell = ws.cell(row=i + currentXlsxRow,
                       column=col_index + currentXlsxColumn, value=value)
        if col_index == 0:
            cell.font = boldFont
        cell.border = thinBorder

    # rowData = []
    # rowData.append("")
    # rowData.append("")
    # rowData.append("")
    # rowData.append("CELKEM")
    # for q in qResult.fetchall():
    #     for c in q:
    #         rowData.append(c)
    # ws.append(rowData)


def PrintOutDemoResult(conn, cursor, totalOblec, totalBoty, totalKosme, totalKabel, sqlQueryJoinCommonCZ, goodsTypeStr):
    obleceniQuery = (
        f"SELECT * FROM {sqlQueryJoinCommonCZ} AND {goodsTypeStr} LIKE '%obleč%';"
    )
    botyQueryyyyyy = (
        f"SELECT * FROM {sqlQueryJoinCommonCZ} AND {goodsTypeStr} LIKE '%boty%';"
    )
    kosmetikaQuery = (
        f"SELECT * FROM {sqlQueryJoinCommonCZ} AND {goodsTypeStr} LIKE '%kosmetika%';"
    )
    kabelkyQuery = (
        f"SELECT * FROM {sqlQueryJoinCommonCZ} AND {goodsTypeStr} LIKE '%kabel%';"
    )

    obleceni = cursor.execute(obleceniQuery)
    for o in obleceni.fetchall():
        totalOblec = totalOblec + o[10]

    boty = cursor.execute(botyQueryyyyyy)
    for o in boty.fetchall():
        totalBoty = totalBoty + o[10]

    kosmetika = cursor.execute(kosmetikaQuery)
    for o in kosmetika.fetchall():
        totalKosme = totalKosme + o[10]

    kabelky = cursor.execute(kabelkyQuery)
    for o in kabelky.fetchall():
        totalKabel = totalKabel + o[10]

    totalOblec = totalOblec * GgoodsCoefficients["Obleceni"]["plast"]
    totalBoty = totalBoty * GgoodsCoefficients["Boty"]["papir"]
    totalKosme = totalKosme * GgoodsCoefficients["Kosmetika"]["plast"]
    totalKabel = totalKabel * GgoodsCoefficients["Kabelky"]["plast"]

    print("\n=====================================================================")

    obleceniHotovo = f"Výsledek OBLEČENÍ je: {totalOblec} tun."
    print(obleceniHotovo)
    botyHotovo = f"Výsledek BOTY je: {totalBoty} tun."
    print(botyHotovo)
    kosmetikaHotovo = f"Výsledek KOSMETIKA je: {totalKosme} tun."
    print(kosmetikaHotovo)
    kabelkyHotovo = f"Výsledek KABELKY je: {totalKabel} tun."
    print(kabelkyHotovo)

    with open("Výsledky.txt", "w", encoding='utf8') as output:
        output.write(obleceniHotovo)
        output.write(botyHotovo)
        output.write(kosmetikaHotovo)
        output.write(kabelkyHotovo)


def main():

    args = sys.argv
    print(len(args))

    testPath = "Q1_25_M_Final.csv"
    directory = os.getcwd()

    if len(args) != 3:
        csvData = os.path.join(directory, testPath)
    elif len(csvData) == 0:
        sys.exit(2)

    buildDB(csvData, "dodavatele2.csv")

    return 0


if __name__ == "__main__":
    main()
